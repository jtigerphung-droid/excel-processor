import streamlit as st
import pandas as pd
import io
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import zipfile  # Kế thừa phục vụ đóng gói nhiều file hóa đơn ở Bước 3

# --- CẤU HÌNH GIAO DIỆN HỆ THỐNG ---
st.set_page_config(page_title="Hệ thống Phát Hành Sách 2 Giai Đoạn", layout="wide")
st.title("📊 HỆ THỐNG XỬ LÝ DỮ LIỆU PHÁT HÀNH SÁCH")
st.write("Phiên bản CODE1_V3 - Hoàn thiện trọn gói 3 Giai đoạn (Tích hợp xuất File Mẫu Hóa Đơn).")

# --- HÀM TRANG TRÍ EXCEL THEO QUY CHUẨN KẾ TOÁN ---
def trang_tri_sheet(worksheet, tieude_color, has_vat_summary=False, total_row_type="standard"):
    font_tieude = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    fill_tieude = PatternFill(start_color=tieude_color, end_color=tieude_color, fill_type="solid")
    font_noidung = Font(name="Arial", size=10)
    font_tongket = Font(name="Arial", size=10, bold=True)
    
    vien_mong = Side(border_style="thin", color="D9D9D9")
    border_o = Border(left=vien_mong, right=vien_mong, top=vien_mong, bottom=vien_mong)
    
    worksheet.row_dimensions[1].height = 25
    for col_idx in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.font = font_tieude
        cell.fill = fill_tieude
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_o

    max_r = worksheet.max_row
    for row_idx in range(2, max_r + 1):
        worksheet.row_dimensions[row_idx].height = 19
        
        is_total_row = False
        if total_row_type == "grand" and row_idx == max_r:
            is_total_row = True
        elif total_row_type == "standard" and has_vat_summary and row_idx >= max_r - 2:
            is_total_row = True
        elif total_row_type == "standard" and not has_vat_summary and row_idx == max_r:
            is_total_row = True

        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.border = border_o
            
            if is_total_row:
                cell.font = font_tongket
                if total_row_type == "grand":
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            else:
                cell.font = font_noidung

            if col_idx in [1, 3]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if col_idx == 3: cell.number_format = '@'
            elif col_idx == 2:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col_idx == 4:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 6:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if total_row_type == "grand":
                    cell.number_format = "#,##0"
                else:
                    cell.number_format = "0.00"
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0"

    for col in worksheet.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)


# --- ĐIỀU HƯỚNG TÍNH NĂNG QUA TAB ---
tab_giai_doan_1, tab_giai_doan_2, tab_giai_doan_3 = st.tabs([
    "🔄 GIAI ĐOẠN 1: Gộp & Làm Sạch", 
    "🧮 GIAI ĐOẠN 2: Tính Toán Phân Tách PAB21",
    "📝 GIAI ĐOẠN 3: Xuất File Mẫu Hóa Đơn (.bas Mapping)"
])

# ==========================================================================================
# GIAI ĐOẠN 1: BẢO LƯU NGUYÊN BẢN 100% CODE GỐC CHẠY HOÀN CHỈNH CỦA USER
# ==========================================================================================
with tab_giai_doan_1:
    st.header("Bước 1: Gộp Nhiều Sheet & Làm Sạch Diện Rộng")
    st.info("Mã nguồn phần này được bảo lưu nguyên vẹn theo cấu trúc chạy ổn định trước đó.")
    
    uploaded_file = st.file_uploader("Kéo thả file Excel thô của hệ thống vào đây:", type=["xlsx"], key="file_tho_goc")

    if uploaded_file is not None:
        if st.button("🚀 THỰC HIỆN GỘP VÀ LÀM SẠCH PHIÊN BẢN CHUẨN"):
            try:
                excel_file = pd.ExcelFile(uploaded_file)
                all_cleaned_rows = []
                
                for sheet_name in excel_file.sheet_names:
                    df_raw_sheet = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
                    if df_raw_sheet.empty: continue
                    
                    for idx, row in df_raw_sheet.iterrows():
                        row_vals = row.fillna("").astype(str).str.strip().tolist()
                        
                        if not row_vals or len(row_vals) < 7: continue
                        if any(k in "".join(row_vals) for k in ["Tổng cộng", "Thuế VAT", "Thành tiền", "Người lập", "Thủ kho", "Giám đốc"]): continue
                        if "Mã số" in row_vals or "Tên sách" in row_vals or "STT" in row_vals: continue
                        
                        idx_ma = -1
                        idx_ten = -1
                        
                        for c_i, val in enumerate(row_vals):
                            if val and not " " in val and len(val) >= 4 and idx_ma == -1:
                                if re.match(r'^[A-Za-z0-9\-_.]+$', val): idx_ma = c_i
                            elif val and " " in val and len(val) > 10 and idx_ten == -1:
                                idx_ten = c_i
                        
                        if idx_ma == -1 or idx_ten == -1:
                            if row_vals[0].isdigit() or (row_vals[2] != "" and len(row_vals[2]) >= 4):
                                idx_ma = 2
                                idx_ten = 1
                            else:
                                continue
                                
                        try:
                            base_idx = min(idx_ten, idx_ma)
                            if base_idx > 0 and row_vals[base_idx-1].isdigit():
                                stt_val = row_vals[base_idx-1]
                            else:
                                stt_val = ""
                                
                            ten_sach = row_vals[idx_ten]
                            ma_so = row_vals[idx_ma]
                            dvt = row_vals[idx_ma + 1] if (idx_ma + 1) < len(row_vals) else "Cuốn"
                            
                            num_fields = []
                            for v in row_vals[idx_ma + 2:]:
                                if v == "" or v == "nan": num_fields.append(0)
                                else:
                                    v_clean = v.replace(",", "").replace(".", "").strip()
                                    if v_clean.replace("-", "").isdigit(): num_fields.append(float(v_clean))
                                    else: num_fields.append(0)
                                    
                            while len(num_fields) < 5: num_fields.append(0)
                            
                            gia_bia = num_fields[0]
                            ck = num_fields[1]
                            sl = num_fields[2]
                            don_gia = num_fields[3]
                            thanh_tien = num_fields[4]
                            
                            if sl == 0 or ma_so == "" or ma_so == "nan": continue
                            
                            all_cleaned_rows.append([
                                stt_val, ten_sach, ma_so, dvt, gia_bia, ck, sl, don_gia, thanh_tien
                            ])
                        except Exception:
                            continue

                if not all_cleaned_rows:
                    st.error("❌ LỖI: Không trích xuất được dòng dữ liệu sách hợp lệ nào.")
                    st.stop()
                    
                df_master = pd.DataFrame(all_cleaned_rows, columns=['STT', 'Tên sách', 'Mã số', 'ĐVT', 'Giá bìa', 'CK', 'Số lượng', 'Đơn giá', 'Thành tiền'])
                df_master['STT'] = range(1, len(df_master) + 1)
                
                out_sach = io.BytesIO()
                with pd.ExcelWriter(out_sach, engine='openpyxl') as writer:
                    df_master.to_excel(writer, index=False, sheet_name="Du_Lieu_Sach_100")
                    trang_tri_sheet(writer.sheets["Du_Lieu_Sach_100"], "003366")
                    
                st.success("🎉 GIAI ĐOẠN 1 HOÀN THÀNH: XUẤT FILE 9 CỘT CHUẨN!")
                st.dataframe(df_master, use_container_width=True)
                
                st.download_button(
                    label="📥 TẢI FILE DỮ LIỆU SẠCH (Dùng cho Giai đoạn 2)",
                    data=out_sach.getvalue(),
                    file_name=f"DuLieu_Gop_Va_LamSach_Chuan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"Lỗi hệ thống G1: {str(e)}")

# ==========================================================================================
# GIAI ĐOẠN 2: KIẾN TRÚC MÃ NGUỒN THEO PHƯƠNG PHÁP PAB21 (BẢO LƯU CHUẨN CODE1_V2)
# ==========================================================================================
with tab_giai_doan_2:
    st.header("Bước 2: Phân Tách Thuộc Tính & Tô Màu Tab Sheet PAB21")
    
    file_sach = st.file_uploader("Tải lên file Excel ĐÃ LÀM SẠCH CHUẨN:", type=["xlsx"], key="file_sach_pab21_v2")
    
    if file_sach is not None:
        if st.button("🧮 KHỞI CHẠY KIẾN TRÚC TÍNH TOÁN NÂNG CẤP"):
            try:
                wb_in = openpyxl.load_workbook(file_sach, data_only=True)
                if "Du_Lieu_Sach_100" not in wb_in.sheetnames:
                    st.error("❌ LỖI: File không chứa sheet 'Du_Lieu_Sach_100'.")
                    st.stop()
                ws_in = wb_in["Du_Lieu_Sach_100"]
                
                mảng_bán_dương = []
                mảng_trả_âm = []
                
                for row_idx in range(2, ws_in.max_row + 1):
                    row_vals = [ws_in.cell(row=row_idx, column=c).value for c in range(1, 10)]
                    if not row_vals[2] or row_vals[2] == "None": continue
                    
                    val_sl = float(row_vals[6] or 0)
                    if val_sl == 0: continue
                    
                    formatted_row = [
                        row_vals[0], str(row_vals[1] or ""), str(row_vals[2]), str(row_vals[3] or "Cuốn"),
                        float(row_vals[4] or 0), float(row_vals[5] or 0), val_sl, 
                        float(row_vals[7] or 0), float(row_vals[8] or 0)
                    ]
                    
                    if val_sl > 0: mảng_bán_dương.append(formatted_row)
                    else: mảng_trả_âm.append(formatted_row)

                # --- 1. TH GIÁ BÌA BÁN (DƯƠNG) ---
                dic_gia_bia_ban = {}
                for r in mảng_bán_dương:
                    ma = r[2]
                    if ma not in dic_gia_bia_ban: dic_gia_bia_ban[ma] = {"ten": r[1], "dvt": r[3], "bia": r[4], "sl": 0.0}
                    dic_gia_bia_ban[ma]["sl"] += r[6]
                bảng_gia_bia_ban = [[i+1, v["ten"], k, v["dvt"], v["bia"], 0.0, v["sl"], v["bia"], v["sl"]*v["bia"]] for i, (k, v) in enumerate(dic_gia_bia_ban.items())]

                # --- 2. TH GIÁ BÌA TRẢ (ÂM) ---
                dic_gia_bia_tra = {}
                for r in mảng_trả_âm:
                    ma = r[2]
                    if ma not in dic_gia_bia_tra: dic_gia_bia_tra[ma] = {"ten": r[1], "dvt": r[3], "bia": r[4], "sl": 0.0}
                    dic_gia_bia_tra[ma]["sl"] += r[6]
                bảng_gia_bia_tra = [[i+1, v["ten"], k, v["dvt"], v["bia"], 0.0, v["sl"], v["bia"], v["sl"]*v["bia"]] for i, (k, v) in enumerate(dic_gia_bia_tra.items())]

                # --- 3. TH CHIẾT KHẤU BÁN (DƯƠNG) ---
                dic_ck_ban = {}
                for r in mảng_bán_dương:
                    key = f"{r[2]}_{r[5]}"
                    if key not in dic_ck_ban: dic_ck_ban[key] = {"ten": r[1], "ma": r[2], "dvt": r[3], "bia": r[4], "ck": r[5], "sl": 0.0, "dg": r[7]}
                    dic_ck_ban[key]["sl"] += r[6]
                bảng_ck_ban = [[i+1, v["ten"], v["ma"], v["dvt"], v["bia"], v["ck"], v["sl"], v["dg"], v["sl"]*v["dg"]] for i, (k, v) in enumerate(dic_ck_ban.items())]

                # --- 4. TH CHIẾT KHẤU TRẢ (ÂM) ---
                dic_ck_tra = {}
                for r in mảng_trả_âm:
                    key = f"{r[2]}_{r[5]}"
                    if key not in dic_ck_tra: dic_ck_tra[key] = {"ten": r[1], "ma": r[2], "dvt": r[3], "bia": r[4], "ck": r[5], "sl": 0.0, "dg": r[7]}
                    dic_ck_tra[key]["sl"] += r[6]
                bảng_ck_tra = [[i+1, v["ten"], v["ma"], v["dvt"], v["bia"], v["ck"], v["sl"], v["dg"], v["sl"]*v["dg"]] for i, (k, v) in enumerate(dic_ck_tra.items())]

                # --- ĐÓNG GÓI EXCEL ĐA MÀU SẮC TAB ---
                out_report = io.BytesIO()
                cols = ['STT', 'Tên sách', 'Mã số', 'ĐVT', 'Giá bìa', 'CK', 'Số lượng', 'Đơn giá', 'Thành tiền']
                
                with pd.ExcelWriter(out_report, engine='openpyxl') as writer:
                    summary_data = []
                    
                    def ghi_sheet(data, name, h_color, tab_color, has_vat=False):
                        df = pd.DataFrame(data, columns=cols)
                        df.to_excel(writer, index=False, sheet_name=name)
                        ws = writer.sheets[name]
                        ws.sheet_properties.tabColor = tab_color
                        
                        r_last = len(df) + 2
                        s_sl, s_tt = df['Số lượng'].sum(), df['Thành tiền'].sum()
                        ws.cell(row=r_last, column=6).value = "Tổng cộng:"
                        ws.cell(row=r_last, column=7).value = s_sl
                        ws.cell(row=r_last, column=9).value = s_tt
                        
                        v_val, st_val = 0.0, 0.0
                        if has_vat:
                            v_val = s_tt * 0.05
                            st_val = s_tt + v_val
                            ws.cell(row=r_last+1, column=8).value = "VAT 5%:"
                            ws.cell(row=r_last+1, column=9).value = v_val
                            ws.cell(row=r_last+2, column=8).value = "Sau Thuế:"
                            ws.cell(row=r_last+2, column=9).value = st_val
                        
                        trang_tri_sheet(ws, h_color, has_vat_summary=has_vat)
                        return len(df), s_sl, s_tt, v_val, st_val

                    ghi_sheet(mảng_bán_dương, "Du_Lieu_Ban_Duong", "595959", "595959")
                    ghi_sheet(mảng_trả_âm, "Du_Lieu_Tra_Am", "595959", "3B3838")
                    
                    l, q, a, _, _ = ghi_sheet(bảng_gia_bia_ban, "TH_Hang_Ban_Gia_Bia", "002060", "002060")
                    summary_data.append(["TH Hàng Bán Giá Bìa", "Dương", l, q, a, 0, a, "Giá Bìa Gốc"])
                    
                    l, q, a, _, _ = ghi_sheet(bảng_gia_bia_tra, "TH_Hang_Tra_Gia_Bia", "800000", "800000")
                    summary_data.append(["TH Hàng Trả Giá Bìa", "Âm", l, q, a, 0, a, "Bảo toàn dấu âm"])

                    l_b, q_b, a_b, v_b, s_b = ghi_sheet(bảng_ck_ban, "TH_Hang_Ban_Chiet_Khau", "339933", "339933", has_vat=True)
                    summary_data.append(["TH Hàng Bán Chiết Khấu", "Dương", l_b, q_b, a_b, v_b, s_b, "Bán thực tế"])

                    l_t, q_t, a_t, v_t, s_t = ghi_sheet(bảng_ck_tra, "Tong_Hop_Hang_Tra", "C00000", "C00000", has_vat=True)
                    summary_data.append(["Tổng Hợp Hàng Trả", "Âm", l_t, q_t, a_t, v_t, s_t, "Trả thực tế (Âm)"])

                    if bảng_ck_ban:
                        for i, start in enumerate(range(0, len(bảng_ck_ban), 1000), 1):
                            batch = [list(r) for r in bảng_ck_ban[start:start+1000]]
                            for idx, row in enumerate(batch, 1): row[0] = idx
                            hd_name = f"HD {i}"
                            l, q, a, v, s = ghi_sheet(batch, hd_name, "008080", "008080", has_vat=True)
                            summary_data.append([hd_name, "Tách HĐ", l, q, a, v, s, f"Từ dòng {start+1}"])

                    df_sum = pd.DataFrame(summary_data, columns=["Tên Sheet / Hạng mục", "Loại", "Số dòng", "Số lượng", "Trước Thuế", "VAT (5%)", "Sau Thuế", "Ghi chú"])
                    df_sum.to_excel(writer, index=False, sheet_name="Tong_Ket_Chung")
                    ws_sum = writer.sheets["Tong_Ket_Chung"]
                    ws_sum.sheet_properties.tabColor = "1F1F1F"
                    
                    r_net = len(df_sum) + 2
                    ws_sum.cell(row=r_net, column=1).value = "DOANH THU THỰC TẾ (BÁN + TRẢ)"
                    ws_sum.cell(row=r_net, column=4).value = q_b + q_t
                    ws_sum.cell(row=r_net, column=5).value = a_b + a_t
                    ws_sum.cell(row=r_net, column=6).value = v_b + v_t
                    ws_sum.cell(row=r_net, column=7).value = s_b + s_t
                    trang_tri_sheet(ws_sum, "1F1F1F", total_row_type="grand")
                    
                    wb = writer.book
                    wb._sheets = [wb._sheets[wb.sheetnames.index(n)] for n in (["Tong_Ket_Chung"] + [sn for sn in wb.sheetnames if sn != "Tong_Ket_Chung"])]

                st.success("🎉 GIAI ĐOẠN 2 HOÀN TẤT CHÍNH XÁC!")
                st.download_button(label="📥 TẢI MASTER REPORT G2", data=out_report.getvalue(), file_name=f"Master_PAB21_{datetime.now().strftime('%m%d_%H%M')}.xlsx")
            except Exception as e:
                st.error(f"Lỗi G2: {str(e)}")

# ==========================================================================================
# GIAI ĐOẠN 3: XUẤT DỮ LIỆU SANG FILE MẪU HÓA ĐƠN - KẾ THỪA ÁNH XẠ .BAS VÀ ĐÓNG GÓI ZIP
# ==========================================================================================
with tab_giai_doan_3:
    st.header("Bước 3: Trích Xuất Dữ Liệu Sang File Mẫu (.bas Mapping)")
    st.info("Hệ thống sẽ quét toàn bộ các hóa đơn 'HD 1', 'HD 2'... từ file kết quả và bốc dữ liệu thả vào File Mẫu.")
    
    # Cổng nạp 1: File kết quả Master Report từ Bước 2
    g3_result_file = st.file_uploader("1. Tải lên file Kết quả Master (Có chứa các sheet HD 1, HD 2...):", type=["xlsx"], key="g3_res_file")
    # Cổng nạp 2: File Mẫu trắng (Template Excel để up hóa đơn điện tử)
    g3_template_file = st.file_uploader("2. Tải lên FILE MẪU TRẮNG (Template Excel):", type=["xlsx", "xlsm"], key="g3_tpl_file")
    
    if g3_result_file is not None and g3_template_file is not None:
        if st.button("📝 TIẾN HÀNH TRÍCH XUẤT HÓA ĐƠN ĐIỆN TỬ"):
            try:
                # Đọc file kết quả bằng openpyxl dạng data_only để lấy trị số thô
                wb_res = openpyxl.load_workbook(g3_result_file, data_only=True)
                
                # Tạo một bộ nhớ đệm để chứa file ZIP xuất ra cho người dùng tải
                zip_buffer = io.BytesIO()
                
                # Đếm xem có bao nhiêu hóa đơn được trích xuất thành công
                success_count = 0
                
                with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                    
                    # Duyệt qua các sheet trong file kết quả 
                    for sheet_name in wb_res.sheetnames:
                        if "HD " in sheet_name: [cite: 3]
                            ws_res = wb_res[sheet_name]
                            
                            last_row = ws_res.max_row
                            # Xác định dòng kết thúc hàng hóa theo thuật toán (lastRowI - 3) 
                            data_end_row = last_row - 3 [cite: 5]
                            num_data_rows = data_end_row - 1 # Bỏ đi dòng tiêu đề [cite: 6]
                            
                            if num_data_rows < 1:
                                continue
                                
                            # Đọc trị trị thanh toán sau thuế đã được làm tròn ở dòng sát cuối (lastRowI - 1) [cite: 4]
                            val_ae11 = round(float(ws_res.cell(row=last_row - 1, column=9).value or 0), 0) [cite: 4]
                            
                            # Đọc mảng thô dữ liệu hàng hóa từ dòng 2 đến dòng kết thúc dữ liệu
                            rows_data = []
                            for r_idx in range(2, data_end_row + 1):
                                r_vals = [ws_res.cell(row=r_idx, column=c_idx).value for c_idx in range(1, 10)]
                                rows_data.append(r_vals)
                                
                            # NẠP FILE MẪU TRẮNG LÊN BỘ NHỚ ĐỂ ĐIỀN DỮ LIỆU [cite: 7]
                            # Đọc không dùng data_only để giữ nguyên công thức sẵn có của File mẫu
                            g3_template_file.seek(0)
                            wb_tpl = openpyxl.load_workbook(g3_template_file, data_only=False)
                            ws_tpl = wb_tpl.worksheets[0] # Lấy sheet đầu tiên của file mẫu [cite: 7]
                            
                            # --- THỰC THI ÁNH XẠ TOÀN DIỆN CỦA FILE .BAS --- 
                            for offset, row_items in enumerate(rows_data):
                                target_r = 11 + offset # Bắt đầu ghi từ dòng 11 trở đi 
                                
                                # Trích xuất thông tin theo thứ tự cột từ mảng kết quả
                                t_ten_sach  = row_items[1] # Cột B: Tên sách 
                                t_ma_so     = row_items[2] # Cột C: Mã số 
                                t_dvt       = row_items[3] # Cột D: ĐVT [cite: 9]
                                t_gia_bia   = row_items[4] # Cột E: Giá bìa 
                                t_ck        = row_items[5] # Cột F: CK 
                                t_sl        = row_items[6] # Cột G: Số lượng 
                                t_don_gia   = row_items[7] # Cột H: Đơn giá bán 
                                t_thanh_tien= row_items[8] # Cột I: Thành tiền chưa thuế [cite: 12, 13]
                                
                                # Ánh xạ chuẩn xác vào File Mẫu tương tự cơ chế sao chép Range: 
                                ws_tpl.cell(row=target_r, column=22).value = t_ten_sach  # Cột V 
                                ws_tpl.cell(row=target_r, column=21).value = t_ma_so     # Cột U 
                                ws_tpl.cell(row=target_r, column=26).value = t_dvt       # Cột Z 
                                ws_tpl.cell(row=target_r, column=23).value = t_gia_bia   # Cột W 
                                ws_tpl.cell(row=target_r, column=24).value = t_ck        # Cột X 
                                ws_tpl.cell(row=target_r, column=27).value = t_sl        # Cột AA 
                                ws_tpl.cell(row=target_r, column=28).value = t_don_gia   # Cột AB 
                                ws_tpl.cell(row=target_r, column=29).value = t_thanh_tien# Cột AC [cite: 12, 13]
                                
                                # Khôi phục cơ chế STT cột A tịnh tiến tuần tự 
                                ws_tpl.cell(row=target_r, column=1).value = offset + 1 # Cột A [cite: 19]
                            
                            # Xử lý quy tắc sao chép (AutoFill) từ ô B11 của hệ thống cũ 
                            val_b11 = ws_tpl.cell(row=11, column=2).value [cite: 16]
                            if num_data_rows > 1: [cite: 16]
                                for r_fill in range(12, 11 + num_data_rows):
                                    ws_tpl.cell(row=r_fill, column=2).value = val_b11 [cite: 16]
                                    
                            # Gán tổng tiền đã được làm tròn vào ô AE11 tĩnh [cite: 17, 18]
                            ws_tpl.cell(row=11, column=31).value = val_ae11 # Cột AE = Cột 31 [cite: 18]
                            
                            # Ghi dữ liệu của hóa đơn cụ thể này vào bộ nhớ tạm
                            single_invoice_buffer = io.BytesIO()
                            wb_tpl.save(single_invoice_buffer)
                            
                            # Đặt tên cho file hóa đơn mới xuất ra đúng định dạng 
                            out_filename = f"Up_Hoa_Don_{sheet_name}_{datetime.now().strftime('%H%m')}.xlsx" [cite: 21]
                            
                            # Thêm file đơn lẻ này vào gói nén ZIP
                            zip_file.writestr(out_filename, single_invoice_buffer.getvalue())
                            success_count += 1
                
                if success_count == 0:
                    st.warning("⚠️ Không tìm thấy sheet 'HD ' hợp lệ nào có chứa dữ liệu để xuất.")
                    st.stop()
                    
                st.success(f"🎉 XUẤT HOÀN TẤT KHỚP CHUẨN: Đã đóng gói {success_count} file hóa đơn điện tử mẫu dạng nén!")
                
                # Tạo nút tải gói ZIP chứa đầy đủ hóa đơn về máy tính
                st.download_button(
                    label="📥 TẢI XUỐNG TOÀN BỘ FILE MẪU HÓA ĐƠN (.ZIP)",
                    data=zip_buffer.getvalue(),
                    file_name=f"Bo_Hoa_Don_Up_He_Thong_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                    mime="application/zip"
                )
                
            except Exception as e:
                st.error(f"Lỗi nghiêm trọng tại Giai đoạn 3: {str(e)}")
